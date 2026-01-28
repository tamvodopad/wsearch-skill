#!/usr/bin/env python3
"""
Создание Excel файла с результатами поиска участников ВОВ.
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Установите openpyxl: pip3 install openpyxl")
    sys.exit(1)


def create_results_excel(
    confirmed: list,
    candidates: list,
    search_log: list,
    variants: list,
    output_path: str = "results.xlsx"
):
    """
    Создаёт Excel файл с 4 листами.

    Args:
        confirmed: Список подтверждённых записей
        candidates: Список кандидатов
        search_log: Журнал поиска
        variants: Использованные варианты написания
        output_path: Путь к выходному файлу
    """
    wb = Workbook()

    # Стили
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # === Лист 1: Итог ===
    ws1 = wb.active
    ws1.title = "Итог"

    headers1 = [
        "ФИО", "Год рождения", "Статус", "Ссылка Память народа",
        "Основание", "Источник географии"
    ]
    ws1.append(headers1)

    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row in confirmed:
        ws1.append([
            row.get('name', ''),
            row.get('birth_year', ''),
            row.get('status', ''),
            row.get('pamyat_url', ''),
            row.get('confirmation_level', ''),
            row.get('geography_source', '')
        ])

    # Автоширина колонок
    for col in range(1, len(headers1) + 1):
        ws1.column_dimensions[get_column_letter(col)].width = 20
    ws1.column_dimensions['A'].width = 35  # ФИО шире
    ws1.column_dimensions['D'].width = 50  # URL Память народа

    # === Лист 2: Кандидаты ===
    ws2 = wb.create_sheet("Кандидаты")

    headers2 = ["ФИО", "Год рождения", "Причина сомнения", "Ссылка", "География"]
    ws2.append(headers2)

    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    for row in candidates:
        ws2.append([
            row.get('name', ''),
            row.get('birth_year', ''),
            row.get('doubt_reason', ''),
            row.get('url', ''),
            row.get('geography_notes', '')
        ])

    for col in range(1, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 25
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['D'].width = 50

    # === Лист 3: Журнал ===
    ws3 = wb.create_sheet("Журнал")

    headers3 = ["Вариант запроса", "Попытка", "Сигнал", "Количество результатов"]
    ws3.append(headers3)

    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col)
        cell.font = header_font_white
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    for row in search_log:
        ws3.append([
            row.get('query', ''),
            row.get('attempt', ''),
            row.get('signal', ''),
            row.get('count', '')
        ])

    ws3.column_dimensions['A'].width = 40
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 12
    ws3.column_dimensions['D'].width = 25

    # === Лист 4: Варианты ===
    ws4 = wb.create_sheet("Варианты")

    headers4 = ["Вариант написания"]
    ws4.append(headers4)

    cell = ws4.cell(row=1, column=1)
    cell.font = header_font_white
    cell.fill = PatternFill(start_color="9E480E", end_color="9E480E", fill_type="solid")
    cell.border = thin_border

    for variant in variants:
        ws4.append([variant])

    ws4.column_dimensions['A'].width = 50

    # Сохранение
    wb.save(output_path)
    print(f"Создан файл: {output_path}")
    print(f"  - Итог: {len(confirmed)} записей")
    print(f"  - Кандидаты: {len(candidates)} записей")
    print(f"  - Журнал: {len(search_log)} записей")
    print(f"  - Варианты: {len(variants)} шт")


def normalize_status(raw_status: str) -> str:
    """Нормализует статус солдата."""
    status_lower = raw_status.lower().strip()

    if any(x in status_lower for x in ['погиб', 'убит']):
        return 'Погиб'
    elif any(x in status_lower for x in ['пропал без вести', 'пропал б/в', 'пропал']):
        return 'Пропал без вести'
    elif 'умер от ран' in status_lower:
        return 'Умер от ран'
    elif 'умер в плену' in status_lower:
        return 'Умер в плену'
    elif any(x in status_lower for x in ['вернулся', 'жив', 'демобилизован']):
        return 'Вернулся'
    else:
        return raw_status


if __name__ == "__main__":
    # Пример использования
    if len(sys.argv) > 1:
        # Загрузка данных из JSON файла
        with open(sys.argv[1], 'r', encoding='utf-8') as f:
            data = json.load(f)

        create_results_excel(
            confirmed=data.get('confirmed', []),
            candidates=data.get('candidates', []),
            search_log=data.get('search_log', []),
            variants=data.get('variants', []),
            output_path=data.get('output_path', 'results.xlsx')
        )
    else:
        # Демо с тестовыми данными
        demo_confirmed = [
            {
                'name': 'Иванов Иван Иванович',
                'birth_year': '1920',
                'status': 'Погиб',
                'pamyat_url': 'https://pamyat-naroda.ru/heroes/123',
                'confirmation_level': 'A',
                'geography_source': 'Место рождения'
            }
        ]

        demo_candidates = [
            {
                'name': 'Петров Пётр Петрович',
                'birth_year': '1918',
                'doubt_reason': 'Неточное совпадение района',
                'url': 'https://warsearch.ru/card/456',
                'geography_notes': 'Можгинский р-н (без уточнения деревни)'
            }
        ]

        demo_log = [
            {'query': 'Вишур', 'attempt': 1, 'signal': 'A', 'count': 47},
            {'query': 'Вишур Можгинский', 'attempt': 1, 'signal': 'A', 'count': 32}
        ]

        demo_variants = ['Вишур', 'Вишур Можгинский район', 'д. Вишур']

        create_results_excel(
            demo_confirmed, demo_candidates, demo_log, demo_variants,
            'demo_results.xlsx'
        )
