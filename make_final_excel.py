#!/usr/bin/env python3
"""Создание финального Excel с солдатами из Вишура и их статусами."""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Загружаем подтверждённые записи со статусами
with open("/Users/popov/Downloads/confirmed_with_status.json", "r", encoding="utf-8") as f:
    confirmed = json.load(f)

# Загружаем кандидатов
with open("/Users/popov/Downloads/candidates.json", "r", encoding="utf-8") as f:
    candidates = json.load(f)

# Создаём Excel
wb = Workbook()

# Стили
header_fill_blue = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_fill_orange = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Цвета для статусов
status_colors = {
    'Погиб': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    'Погиб/Пропал': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    'Умер от ран': PatternFill(start_color="FF8C8C", end_color="FF8C8C", fill_type="solid"),
    'Умер в плену': PatternFill(start_color="FF8C8C", end_color="FF8C8C", fill_type="solid"),
    'Пропал без вести': PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    'Плен': PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
    'Ранен': PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),
    'Награждён': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
    'Вернулся': PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid"),
    'Неизвестен': PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"),
}

def add_confirmed_sheet(wb, title, data, header_fill, is_first=False):
    """Добавляет лист с подтверждёнными записями."""
    if is_first:
        ws = wb.active
        ws.title = title
    else:
        ws = wb.create_sheet(title)

    # Заголовки
    headers = ["№", "ФИО", "Год", "Место рождения", "Место службы", "Статус", "Ссылка", "Уровень"]
    ws.append(headers)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Данные
    for i, r in enumerate(data, 1):
        fio = f"{r['f']} {r['n']} {r['p']}"
        status = r.get('status', '')
        row_num = i + 1
        ws.append([i, fio, r['y'], r['b'], r['s'], status, "", "A"])

        # Цвет статуса
        status_cell = ws.cell(row=row_num, column=6)
        if status in status_colors:
            status_cell.fill = status_colors[status]

        # Кликабельная ссылка
        link_cell = ws.cell(row=row_num, column=7)
        if r.get('u'):
            link_cell.hyperlink = r['u']
            link_cell.value = "Ссылка"
            link_cell.style = "Hyperlink"

    # Ширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8

    return ws

def add_candidates_sheet(wb, title, data, header_fill):
    """Добавляет лист с кандидатами."""
    ws = wb.create_sheet(title)

    # Заголовки
    headers = ["№", "ФИО", "Год", "Место рождения", "Место службы", "Статус", "Ссылка", "Уровень"]
    ws.append(headers)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    # Данные
    for i, r in enumerate(data, 1):
        fio = f"{r['f']} {r['n']} {r['p']}"
        row_num = i + 1
        ws.append([i, fio, r['y'], r['b'], r['s'], "", "", "C"])

        # Кликабельная ссылка
        link_cell = ws.cell(row=row_num, column=7)
        if r.get('u'):
            link_cell.hyperlink = r['u']
            link_cell.value = "Ссылка"
            link_cell.style = "Hyperlink"

    # Ширина колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 8

    return ws

# Создаём листы
add_confirmed_sheet(wb, "Итог", confirmed, header_fill_blue, is_first=True)
add_candidates_sheet(wb, "Кандидаты", candidates, header_fill_orange)

# Лист со статистикой
ws_stats = wb.create_sheet("Статистика")
ws_stats.append(["Статус", "Количество"])
ws_stats.cell(row=1, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws_stats.cell(row=1, column=1).font = header_font
ws_stats.cell(row=1, column=2).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws_stats.cell(row=1, column=2).font = header_font

# Подсчёт статистики
stats = {}
for r in confirmed:
    s = r.get('status', 'Неизвестен')
    stats[s] = stats.get(s, 0) + 1

for status, count in sorted(stats.items(), key=lambda x: -x[1]):
    row = ws_stats.max_row + 1
    ws_stats.append([status, count])
    if status in status_colors:
        ws_stats.cell(row=row, column=1).fill = status_colors[status]

ws_stats.column_dimensions['A'].width = 25
ws_stats.column_dimensions['B'].width = 15

# Лист с вариантами
ws_var = wb.create_sheet("Варианты")
ws_var.append(["Вариант написания"])
ws_var.cell(row=1, column=1).fill = PatternFill(start_color="9E480E", end_color="9E480E", fill_type="solid")
ws_var.cell(row=1, column=1).font = header_font
for v in ["Вишур", "Вишур Кизнерский", "Вишурка", "Вичурка"]:
    ws_var.append([v])
ws_var.column_dimensions['A'].width = 30

# Сохранение
output = "/Users/popov/Downloads/vishur_soldiers_final.xlsx"
wb.save(output)
print(f"Создан файл: {output}")
print(f"  - Итог (подтверждённые): {len(confirmed)} записей")
print(f"  - Кандидаты: {len(candidates)} записей")
print(f"\nСтатистика статусов:")
for status, count in sorted(stats.items(), key=lambda x: -x[1]):
    print(f"  {status}: {count}")
